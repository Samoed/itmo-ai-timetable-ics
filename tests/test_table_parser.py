from datetime import datetime
from pathlib import Path

import openpyxl
import pytest
from dateutil import tz
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from itmo_ai_timetable_ics.schemes import Pair
from itmo_ai_timetable_ics.timetable_file import ScheduleParser

# Set up dates and time slots
timezone = tz.gettz("Europe/Moscow")
TOTAL_DAYS = 3
OFFSET_HEAD = 7
times = ["10:00-11:30", "11:40-13:10", "13:30-15:00", "15:20-16:50", "17:00-18:30", "18:40-20:10"]

DATE_COLUMN = 2
DAY_LENGTH = len(times) + 1


@pytest.fixture
def sample_workbook() -> Workbook:
    wb = Workbook()
    sheet = wb.active

    # Set up header
    sheet["A1"] = "Добро пожаловать в расписание 2го семестра в AI Talent Hub"
    sheet.merge_cells("A1:K1")

    # Set up February section
    sheet["A5"] = "Февраль"
    sheet.merge_cells("A5:K5")
    sheet["B6"] = "5 - 10 февраля (1 неделя)"
    sheet.merge_cells("B6:K6")

    now = datetime.now(tz=timezone)
    for i in range(TOTAL_DAYS):
        sheet.cell(
            row=OFFSET_HEAD + i * DAY_LENGTH,
            column=DATE_COLUMN,
            value=datetime(now.year, now.month, 5 + i),  # noqa: DTZ001
        )
        sheet.merge_cells(
            start_row=OFFSET_HEAD + i * DAY_LENGTH,
            start_column=DATE_COLUMN,
            end_row=OFFSET_HEAD + (i + 1) * DAY_LENGTH - 1,
            end_column=DATE_COLUMN,
        )
        for j, time in enumerate(times):
            sheet.cell(row=7 + i * 7 + j, column=4, value=j + 1)
            sheet.cell(row=7 + i * 7 + j, column=5, value=time)

    # Set up classes
    sheet["F11"] = "Безопасность ИИ\nЧат курса"
    sheet["F11"].fill = PatternFill(start_color="FFD9D2E9", end_color="FFD9D2E9", fill_type="solid")
    sheet["G11"] = "Машинное обучение на больших данных (Big Data ML)\nЧат курса"
    sheet["G11"].fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
    sheet["F12"] = "Разработка приложений разговорного искусственного интеллекта"
    sheet["F12"].fill = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")

    sheet["F14"] = "Специальные главы биоинформатики"
    sheet["F14"].fill = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
    sheet.merge_cells("F14:F15")
    sheet["F18"] = "Глубокие генеративные модели (Deep Generative Models)\nЧат курса"
    sheet["F18"].fill = PatternFill(start_color="FFE6E0EC", end_color="FFE6E0EC", fill_type="solid")
    return wb


@pytest.fixture
def timetable_file(tmp_path: Path, sample_workbook: Workbook) -> ScheduleParser:
    file_path = tmp_path / "test_timetable.xlsx"
    # file_path = "test_timetable.xlsx"
    sample_workbook.save(file_path)
    return ScheduleParser(str(file_path), 0)


def test_init(timetable_file: ScheduleParser):
    assert isinstance(timetable_file.sheet, openpyxl.worksheet.worksheet.Worksheet)


def test_get_days(timetable_file: ScheduleParser):
    days = timetable_file._get_days()
    assert len(days) == TOTAL_DAYS
    assert all(isinstance(day, openpyxl.worksheet.merge.MergedCellRange) for day in days)


def test_pair_time(timetable_file: ScheduleParser):
    cell = timetable_file.sheet["E7"]
    day = datetime(2024, 2, 5, tzinfo=timezone)
    start, end = timetable_file.pair_time(cell, day)
    assert start == datetime(day.year, day.month, day.day, 10, 0, tzinfo=timezone)
    assert end == datetime(day.year, day.month, day.day, 11, 30, tzinfo=timezone)


def test_find_time_in_cell(timetable_file: ScheduleParser):
    cell_value = "Машинное обучение на больших данных (Big Data ML)\nЧат курса"
    cleaned, start, end = timetable_file.find_time_in_cell(cell_value)
    assert cleaned == "Машинное обучение на больших данных (Big Data ML)\nЧат курса"
    assert start is None
    assert end is None


def test_find_key_words_in_cell(timetable_file: ScheduleParser):
    cell_value = "Глубокие генеративные модели (Deep Generative Models)\nЧат курса"
    cleaned, keyword = timetable_file.find_key_words_in_cell(cell_value)
    assert cleaned == "Глубокие генеративные модели (Deep Generative Models)\nЧат курса"
    assert keyword is None


def test_process_cell(timetable_file: ScheduleParser):
    cell = timetable_file.sheet["F11"]
    title, keyword, link = timetable_file.process_cell(cell)
    assert title == "Безопасность ИИ Чат курса"
    assert keyword is None
    assert link is None


def test_parse(timetable_file: ScheduleParser):
    now = datetime.now(tz=timezone)
    pairs = timetable_file.parse()
    assert len(pairs) == 5
    assert isinstance(pairs[0], Pair)
    assert pairs[0].name == "Безопасность ИИ Чат курса"
    assert pairs[0].start_time == datetime(now.year, now.month, 5, 17, 0, tzinfo=timezone)
    assert pairs[0].end_time == datetime(now.year, now.month, 5, 18, 30, tzinfo=timezone)
