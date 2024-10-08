from collections.abc import Generator, Iterable
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from dateutil import tz
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.worksheet.worksheet import Worksheet

from itmo_ai_timetable.cleaner import course_name_cleaner
from itmo_ai_timetable.logger import get_logger
from itmo_ai_timetable.schemes import Pair
from itmo_ai_timetable.settings import Settings

logger = get_logger(__name__)


class ScheduleParser:
    def __init__(self, path: str, sheet: str) -> None:
        self.settings = Settings()
        self.timezone = tz.gettz(self.settings.tz)
        self.sheet = self._load_workbook(path, sheet)

    def _download_excel(self, url: str) -> Path:
        request = requests.get(url, timeout=5)
        file_path = Path("file.xlsx")
        with file_path.open("wb") as file:
            file.write(request.content)
        return file_path

    def _load_workbook(self, path: str, sheet: str) -> Worksheet:
        logger.info("Open file %s", path)
        if path.startswith("http"):
            path = str(self._download_excel(path))
        workbook = openpyxl.load_workbook(path)
        return workbook[sheet]

    def parse(self) -> list[Pair]:
        logger.info("Start parse")
        pairs = []
        for day_cell in self._get_days():
            day = self._get_day(day_cell)
            if day is None:
                continue
            pairs.extend(self._parse_day(day, day_cell))
        logger.info("End parse")
        return pairs

    def _get_days(self) -> list[MergedCellRange]:
        return [
            cell_range
            for cell_range in self.sheet.merged_cells.ranges
            if cell_range.min_col == self.settings.days_column == cell_range.max_col
        ]

    def _get_day(self, day_cell: MergedCellRange) -> datetime | None:
        day = self._get_first_cell_from_range(day_cell).value
        if day is None:
            return None
        if not isinstance(day, datetime):
            raise ValueError(f"Day should be datetime, got {type(day)} in cell range {day_cell} with value {day}")
        day = day.astimezone(self.timezone)
        self._validate_day(day, day_cell)
        return day

    def _validate_day(self, day: datetime, day_cell: MergedCellRange) -> None:
        if (datetime.now(tz=self.timezone) - day).days > self.settings.max_days_difference:
            raise ValueError(f"Date {day.date()} is in previous semester, cell range {day_cell}")

    def _parse_day(self, day: datetime, day_cell: MergedCellRange) -> list[Pair]:
        pairs = []
        for row in self._iter_day_rows(day_cell):
            if row[0].value is None:
                continue
            # first cell in row is time, other cells are pairs
            pair_start, pair_end = self._get_pair_time(row[0], day)
            pairs.extend(self._parse_row(row[1:], pair_start, pair_end))
        return pairs

    def _iter_day_rows(self, day_cell: MergedCellRange) -> Generator[tuple[Cell, ...], None, None]:
        if day_cell.min_col is None or day_cell.max_col is None:
            raise ValueError(f"Day cell range should have min_row and max_row, got {day_cell}")
        return self.sheet.iter_rows(
            min_row=day_cell.min_row,
            max_row=day_cell.max_row,
            min_col=day_cell.min_col + self.settings.timetable_offset,
            max_col=day_cell.max_col + self.settings.timetable_offset + self.settings.timetable_len,
        )

    def _get_pair_time(self, cell: Cell, day: datetime) -> tuple[datetime, datetime]:
        time = cell.value
        if not isinstance(time, str):
            raise ValueError(f"Time should be string, got {type(time)}")
        start_time, end_time = time.split("-")
        start_hour, start_minute = map(int, start_time.split(":"))
        end_hour, end_minute = map(int, end_time.split(":"))
        return (
            day.replace(hour=start_hour, minute=start_minute).astimezone(self.timezone),
            day.replace(hour=end_hour, minute=end_minute).astimezone(self.timezone),
        )

    def _parse_row(self, cells: Iterable[Cell], pair_start: datetime, pair_end: datetime) -> list[Pair]:
        pairs = []
        for cell in cells:
            if cell.value is None or cell.value == "":
                continue
            title, pair_type, link = self._process_cell(cell)

            title, parsed_pair_start, parsed_pair_end = self._find_time_in_cell(title)
            if parsed_pair_start and parsed_pair_end:
                pair_start = pair_start.replace(hour=parsed_pair_start[0], minute=parsed_pair_start[1])
                pair_end = pair_end.replace(hour=parsed_pair_end[0], minute=parsed_pair_end[1])
            if title and title not in self.settings.courses_to_skip:
                pairs.append(
                    Pair(
                        start_time=pair_start,
                        end_time=pair_end,
                        name=title,
                        pair_type=pair_type,
                        link=link,
                    ),
                )
        return pairs

    def _process_cell(self, cell: Cell | MergedCell) -> tuple[str, str | None, str | None]:
        if cell.value is None:
            return "", None, None

        if isinstance(cell, MergedCell):
            merge_range = self._find_mergedcell_mergerange(cell)
            cell = self._get_first_cell_from_range(merge_range)

        link = cell.hyperlink.target if cell.hyperlink else None
        cell_title = self._clean_cell_value(cell)
        cell_title, key_word = self._find_key_words_in_cell(cell_title)
        cell_title = course_name_cleaner(cell_title)

        return cell_title, key_word, link

    def _find_mergedcell_mergerange(self, merged_cell: MergedCell) -> MergedCellRange:
        for merged_range in self.sheet.merged_cells.ranges:
            if (
                merged_range.min_col <= merged_cell.column <= merged_range.max_col
                and merged_range.min_row <= merged_cell.row <= merged_range.max_row
                and isinstance(merged_range, MergedCellRange)
            ):
                return merged_range
        raise ValueError(f"Can't find merged cell range for cell {merged_cell}")

    def _get_first_cell_from_range(self, cell_range: MergedCellRange) -> Cell:
        for row, col in cell_range.cells:
            return self.sheet.cell(row=row, column=col)
        raise ValueError(f"Can't find cell in merged cell range {cell_range}")

    def _clean_cell_value(self, cell: Cell) -> str:
        if not isinstance(cell.value, str):
            raise ValueError(f"Cell value should be string, got {type(cell.value)}")
        return cell.value.replace("/", "\\").replace("\n", " ").strip()

    def _find_key_words_in_cell(self, cell_title: str) -> tuple[str, str | None]:
        if not isinstance(cell_title, str):
            raise ValueError(f"Cell title should be string, got {type(cell_title)}")
        for key_word in self.settings.keywords:
            if key_word in cell_title:
                return cell_title.replace(key_word, "").strip(), key_word
        return cell_title, None

    def _find_time_in_cell(self, cell: str) -> tuple[str, tuple[int, int] | None, tuple[int, int] | None]:
        """Find time in cell.

        "Публичные выступления 1 / Финансовая грамотность 3 17:00 - 19:15" should be 17:00 and 19:15
        """
        cell = cell.strip()
        if not cell or ("-" not in cell and ":" not in cell):
            return cell, None, None

        start_time = end_time = None
        cell = cell.replace("-", " ")
        for time in cell.split():
            if ":" not in time:
                continue
            try:
                hour, minute = map(int, time.split(":"))
            except ValueError as e:
                raise ValueError(f"Invalid time format in cell: {cell}") from e

            if start_time is None:
                start_time = (hour, minute)
            else:
                end_time = (hour, minute)
            cell = cell.replace(time, "").strip()

        return cell, start_time, end_time
