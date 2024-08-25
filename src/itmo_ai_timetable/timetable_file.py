from datetime import datetime
from typing import ClassVar

import openpyxl
from dateutil import tz
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.merge import MergedCellRange

from .logger import get_logger
from .schemes import Pair

logger = get_logger(__name__)
timezone = tz.gettz("Europe/Moscow")


class ScheduleParser:
    DAYS_COLUMN = 2  # Column B
    TIMETABLE_OFFSET = 3  # Offset between date and timetable
    TIMETABLE_LEN = 5  # Length of timetable

    # Extra words in cell that contains type of pair
    KEYWORDS: ClassVar[list[str]] = [
        "Экзамен",
        "Зачет",
        "Дифф. зачет",
    ]

    def __init__(self, path: str, sheet: int) -> None:
        logger.info("Open file %s", path)
        wb = openpyxl.load_workbook(path)
        self.sheet = wb.worksheets[sheet]

    def _get_days(self) -> list[MergedCellRange]:
        days_cells = []

        for merged_cell_range in self.sheet.merged_cells.ranges:
            if merged_cell_range.min_col == self.DAYS_COLUMN and merged_cell_range.max_col == self.DAYS_COLUMN:
                days_cells.append(merged_cell_range)
        return days_cells

    def find_mergedcell_mergerange(self, merged_cell: MergedCell) -> MergedCellRange:
        for merged_range in self.sheet.merged_cells.ranges:
            if (
                merged_range.min_col <= merged_cell.column <= merged_range.max_col
                and merged_range.min_row <= merged_cell.row <= merged_range.max_row
                and isinstance(merged_range, MergedCellRange)
            ):
                return merged_range
        raise ValueError(f"Can't find merged cell range for cell {merged_cell}")

    def get_first_cell_from_mergedcell_range(self, cell_range: MergedCellRange) -> Cell:
        for cell in cell_range.cells:
            return self.sheet.cell(row=cell[0], column=cell[1])
        raise ValueError(f"Can't find cell in merged cell range {cell_range}")

    def find_mergedcell_bounds(self, merged_cell: MergedCell) -> tuple[int, int, int, int]:
        merged_range = self.find_mergedcell_mergerange(merged_cell)
        if (
            merged_range.min_row is not None
            and merged_range.min_col is not None
            and merged_range.max_row is not None
            and merged_range.max_col is not None
        ):
            return (
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col,
            )
        raise ValueError(f"Can't find bounds for merged cell {merged_cell}")

    def pair_time(self, cell: Cell, day: datetime) -> tuple[datetime, datetime]:
        time = cell.value
        if not isinstance(time, str):
            raise ValueError(f"Time should be string, got {type(time)}")
        start_time, end_time = time.split("-")
        start_hour, start_minute = map(int, start_time.split(":"))
        end_hour, end_minute = map(int, end_time.split(":"))
        pair_start = day.replace(hour=start_hour, minute=start_minute).astimezone(timezone)
        pair_end = day.replace(hour=end_hour, minute=end_minute).astimezone(timezone)
        return pair_start, pair_end

    def find_time_in_cell(self, cell: str) -> tuple[str, tuple[int, int] | None, tuple[int, int] | None]:
        """
        Find time in cell

        "Публичные выступления 1 / Финансовая грамотность 3 17:00 - 19:15" should be 17:00 and 19:15
        """
        cell = cell.strip()
        if cell == "":
            return cell, None, None

        start_time = None
        end_time = None
        cell = cell.replace("-", " ")
        for time in cell.split():
            if time.find(":") == -1:
                continue
            try:
                hour, minute = map(int, time.split(":"))
            except ValueError as e:
                raise e

            if start_time is None:
                start_time = (hour, minute)
            else:
                end_time = (hour, minute)
            cell = cell.replace(time, "").strip()

        return cell, start_time, end_time

    def find_key_words_in_cell(self, cell_title: str) -> tuple[str, str | None]:
        if not isinstance(cell_title, str):
            raise ValueError(f"Cell title should be string, got {type(cell_title)}")
        for key_word in self.KEYWORDS:
            if key_word in cell_title:
                cell_title = cell_title.replace(key_word, "")
                return cell_title, key_word
        return cell_title, None

    def clean_cell_value(self, cell: Cell) -> str:
        if not isinstance(cell.value, str):
            raise ValueError(f"Cell value should be string, got {type(cell.value)}")
        return cell.value.replace("/", "\\").replace("\n", " ").strip()

    def process_cell(self, cell: Cell | MergedCell) -> tuple[str, str | None, str | None]:
        if cell.value is None:
            return "", None, None

        if isinstance(cell, MergedCell):
            merge_range = self.find_mergedcell_mergerange(cell)
            cell = self.get_first_cell_from_mergedcell_range(merge_range)
            link = None if cell.hyperlink is None else cell.hyperlink.target
        elif isinstance(cell, Cell):
            link = None if cell.hyperlink is None else cell.hyperlink.target
        else:
            raise ValueError(f"Unknown cell type {type(cell)}")

        cell_title = self.clean_cell_value(cell)
        cell_title, key_word = self.find_key_words_in_cell(cell_title)

        return cell_title, key_word, link

    def parse(self) -> list[Pair]:
        logger.info("Start parse")
        df = []

        for day_cell in self._get_days():
            day = self.get_first_cell_from_mergedcell_range(day_cell).value
            if not isinstance(day, datetime):
                raise ValueError(f"Day should be datetime, got {type(day)}")
            day = day.astimezone(timezone)
            # sometimes date can be written with incorrect year
            half_year = 180
            if (datetime.now(tz=timezone) - day).days > half_year:
                raise ValueError(f"Date {day.date()} is in previous semester, cell range {day_cell}")

            if day_cell.min_col is None or day_cell.max_col is None:
                raise ValueError(f"Can't find min or max column for cell {day_cell}")

            for row in self.sheet.iter_rows(
                min_row=day_cell.min_row,
                max_row=day_cell.max_row,
                min_col=day_cell.min_col + self.TIMETABLE_OFFSET,
                max_col=day_cell.max_col + self.TIMETABLE_OFFSET + self.TIMETABLE_LEN,
            ):
                if row[0].value is None:
                    continue

                # get time for row
                pair_start, pair_end = self.pair_time(row[0], day)
                for cell in row[1:]:
                    if cell.value is None or cell.value == "":
                        continue
                    title, pair_type, link = self.process_cell(cell)
                    title, parsed_pair_start, parsed_pair_end = self.find_time_in_cell(title)
                    if parsed_pair_start and parsed_pair_end:
                        pair_start = pair_start.replace(hour=parsed_pair_start[0], minute=parsed_pair_start[1])
                        pair_end = pair_end.replace(hour=parsed_pair_end[0], minute=parsed_pair_end[1])
                    if title == "":
                        continue
                    df.append(
                        Pair(
                            start_time=pair_start,
                            end_time=pair_end,
                            name=title,
                            pair_type=pair_type,
                            link=link,
                        )
                    )
        logger.info("End parse")
        return df
