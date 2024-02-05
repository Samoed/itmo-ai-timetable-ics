from datetime import datetime

import openpyxl
import pytz
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.merge import MergedCellRange

from logger import get_logger

logger = get_logger(__name__)
moscow_tz = pytz.timezone("Europe/Moscow")


class TimetableFile:
    DAYS_COLUMN = 2  # Column B
    TIMETABLE_OFFSET = 3  # Offset between date and timetable
    TIMETABLE_LEN = 5  # Length of timetable

    KEYWORDS = [
        "Экзамен",
        "Зачет",
        "Дифф. зачет",
    ]

    def __init__(self, path: str, sheet: int):
        logger.info(f"Open file {path}")
        wb = openpyxl.load_workbook(path)
        self.sheet = wb.worksheets[sheet]

    def _get_days(self) -> list[MergedCellRange]:
        days_cells = []

        for merged_cell_range in self.sheet.merged_cells.ranges:
            if merged_cell_range.min_col == self.DAYS_COLUMN and merged_cell_range.max_col == self.DAYS_COLUMN:
                days_cells.append(merged_cell_range)
        return days_cells

    def find_mergedcell_mergerange(self, merged_cell: MergedCell) -> MergedCellRange:
        for merged_range in self.sheet.merged_cells.ranges:
            if (
                merged_range.min_col <= merged_cell.column <= merged_range.max_col
                and merged_range.min_row <= merged_cell.row <= merged_range.max_row
            ):
                return merged_range

    def get_first_cell_from_mergedcell_range(self, cell_range: MergedCellRange) -> Cell:
        for cell in cell_range.cells:
            return self.sheet.cell(row=cell[0], column=cell[1])

    def find_mergedcell_bounds(self, merged_cell: MergedCell) -> tuple[int, int, int, int]:
        merged_range = self.find_mergedcell_mergerange(merged_cell)
        return (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )

    def pair_time(self, cell: Cell, day: datetime) -> tuple[datetime, datetime]:
        time = cell.value
        start_time, end_time = time.split("-")
        start_hour, start_minute = map(int, start_time.split(":"))
        end_hour, end_minute = map(int, end_time.split(":"))
        pair_start = day.replace(hour=start_hour, minute=start_minute).astimezone(pytz.timezone("Europe/Moscow"))
        pair_end = day.replace(hour=end_hour, minute=end_minute).astimezone(pytz.timezone("Europe/Moscow"))
        return pair_start, pair_end

    def find_time_in_cell(self, cell: str) -> tuple[str, tuple[int, int] | None, tuple[int, int] | None]:
        """
        Find time in cell

        "Публичные выступления 1 / Финансовая грамотность 3 17:00 - 19:15" should be 17:00 and 19:15
        """
        cell = cell.strip()
        if cell == "":
            return cell, None, None

        start_time = None
        end_time = None
        cell = cell.replace("-", " ")
        for time in cell.split(" "):
            if ":" in time:
                try:
                    hour, minute = map(int, time.split(":"))
                except ValueError as e:
                    raise e

                if start_time is None:
                    start_time = (hour, minute)
                else:
                    end_time = (hour, minute)
                cell = cell.replace(time, "")

        return cell.strip(), start_time, end_time

    def find_key_words_in_cell(self, cell_title: str) -> tuple[str, str | None]:
        for key_word in self.KEYWORDS:
            if key_word in cell_title:
                cell_title = cell_title.replace(key_word, "")
                return cell_title, key_word
        return cell_title, None

    def clean_cell_value(self, cell: Cell) -> str:
        return cell.value.replace("/", "\\").replace("\n", " ").strip()

    def process_cell(self, cell: Cell | MergedCell) -> tuple[str, str | None, str | None]:
        if cell.value is None:
            return "", None, None

        if isinstance(cell, MergedCell):
            merge_range = self.find_mergedcell_mergerange(cell)
            cell = self.get_first_cell_from_mergedcell_range(merge_range)
            link = None if cell.hyperlink is None else cell.hyperlink.target
        elif isinstance(cell, Cell):
            link = None if cell.hyperlink is None else cell.hyperlink.target
        else:
            raise ValueError(f"Unknown cell type {type(cell)}")

        cell_title = self.clean_cell_value(cell)
        cell_title, key_word = self.find_key_words_in_cell(cell_title)

        return cell_title, key_word, link

    def parse(self) -> list[tuple[datetime, datetime, str, str | None, str | None]]:
        logger.info("Start parse")
        df = []
        pair_start = None
        pair_end = None
        current_year = datetime.now().year
        for day_cell in self._get_days():
            day = self.get_first_cell_from_mergedcell_range(day_cell).value
            if day.year < current_year:
                raise ValueError(f"Year {day.year} is less than current year {current_year}, date in file {day.date()}, cell range {day_cell}")

            for row in self.sheet.iter_rows(
                min_row=day_cell.min_row,
                max_row=day_cell.max_row,
                min_col=day_cell.min_col + self.TIMETABLE_OFFSET,
                max_col=day_cell.max_col + self.TIMETABLE_OFFSET + self.TIMETABLE_LEN,
            ):
                for cell_num, cell in enumerate(row):
                    if cell.value is None:
                        continue
                    # get time for row
                    if cell_num == 0:
                        pair_start, pair_end = self.pair_time(cell, day)
                        continue
                    title, pair_type, link = self.process_cell(cell)
                    title, parsed_pair_start, parsed_pair_end = self.find_time_in_cell(title)
                    if parsed_pair_start:
                        pair_start = pair_start.replace(hour=parsed_pair_start[0], minute=parsed_pair_start[1])
                        pair_end = pair_end.replace(hour=parsed_pair_end[0], minute=parsed_pair_end[1])
                    if title != "":
                        df.append((pair_start, pair_end, title, pair_type, link))
        logger.info("End parse")
        return df
